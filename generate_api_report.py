import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

results = [
    ('GET', '/', 'API root', 200, 'PASS', ''),
    ('GET', '/health', 'Health check', 200, 'PASS', ''),
    ('POST', '/api/v1/auth/register', 'Register user', 201, 'PASS', ''),
    ('POST', '/api/v1/auth/login', 'Login user', 200, 'PASS', ''),
    ('GET', '/api/v1/auth/me', 'Current user with valid token', 200, 'PASS', ''),
    ('GET', '/api/v1/auth/me', 'Current user with invalid token', 401, 'PASS', ''),
    ('POST', '/api/v1/tickets', 'Create ticket', 201, 'PASS', ''),
    ('GET', '/api/v1/tickets', 'List tickets', 200, 'PASS', ''),
    ('GET', '/api/v1/tickets/{ticket_id}', 'Get ticket', 200, 'PASS', ''),
    ('PATCH', '/api/v1/tickets/{ticket_id}', 'Update ticket', 200, 'PASS', ''),
    ('DELETE', '/api/v1/tickets/{ticket_id}', 'Delete ticket', 204, 'PASS', ''),
    ('GET', '/api/v1/tickets/{ticket_id}/interaction-logs', 'List logs for ticket', 200, 'PASS', ''),
    ('POST', '/api/v1/interaction-logs', 'Create interaction log', 201, 'PASS', ''),
    ('GET', '/api/v1/interaction-logs', 'List interaction logs', 200, 'PASS', ''),
    ('GET', '/api/v1/interaction-logs/{log_id}', 'Get interaction log', 200, 'PASS', ''),
    ('PATCH', '/api/v1/interaction-logs/{log_id}', 'Update interaction log', 200, 'PASS', ''),
    ('DELETE', '/api/v1/interaction-logs/{log_id}', 'Delete interaction log', 204, 'PASS', ''),
    ('GET', '/api/v1/agents', 'List agents reference data', 200, 'PASS', ''),
    ('GET', '/api/v1/stores', 'List stores reference data', 200, 'PASS', ''),
    ('GET', '/api/v1/escalation-rules', 'List escalation rules reference data', 200, 'PASS', ''),
    ('GET', '/api/v1/devices', 'List devices reference data', 200, 'PASS', ''),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'API Test Report'
hf = Font(bold=True, color='FFFFFF', size=11)
hbg = PatternFill('solid', fgColor='2F5496')
pg = PatternFill('solid', fgColor='C6EFCE')
fr = PatternFill('solid', fgColor='FFC7CE')
ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
t = Side(style='thin')
bdr = Border(left=t, right=t, top=t, bottom=t)
for c, h in enumerate(['#', 'Method', 'Endpoint', 'Description', 'Status Code', 'Pass/Fail', 'Reason'], 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf
    cell.fill = hbg
    cell.alignment = ctr
    cell.border = bdr
for row, (m, ep, desc, code, pf, rsn) in enumerate(results, 2):
    bg = pg if pf == 'PASS' else fr
    for c, (v, a) in enumerate(
        zip([row - 1, m, ep, desc, code, pf, rsn], [ctr, ctr, lft, lft, ctr, ctr, lft]), 1
    ):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = bg
        cell.alignment = a
        cell.border = bdr
        if c == 6:
            cell.font = Font(bold=True, color='375623' if pf == 'PASS' else '9C0006')
for i, w in enumerate([5, 10, 42, 32, 12, 12, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
wb.save('api_test_report.xlsx')
print('Saved: api_test_report.xlsx')
